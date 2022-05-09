#!/usr/bin/python3
from openpyxl import load_workbook

wb = load_workbook("QFII重仓流通股.xls")
print("Sheet name is : ", wb.sheetnames)

