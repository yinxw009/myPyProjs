import os
from shutil import move

exl_file = 'QFII重仓流通股.xls'
print(exl_file)
path = r'D:\Projects\PyProjs\excel'
dst_path = r'D:\81_Office\Excel'
i = 0
def move_file(name):
    global i
    for root, lists, files in os.walk(path):
        for file in files:
            if name in file:
                i += 1
                src_path = os.path.join(root, file)
                try:
                    move(src_path, dst_path)
                except:
                    pass

def main():
    from openpyxl import load_workbook
    wb = load_workbook(filename= exl_file)
    ws = wb['待提取候选人']
    for cell in ws["B"]:
        move_file(str(cell.value))
    if i >0:
        print()
    else:
        print()